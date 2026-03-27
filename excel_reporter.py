import os
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import pandas as pd
import io
from PIL import Image

class ExcelReporter:
    def __init__(self, templates_dir="/Users/danielvelasquez/Library/CloudStorage/OneDrive-Personal/Proyectos Incol/Datalogger/Formatos",
                 output_dir="/Users/danielvelasquez/Library/CloudStorage/OneDrive-Personal/Proyectos Incol/Datalogger/Resultados"):
        self.templates_dir = templates_dir
        self.output_dir = output_dir
        
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def generate_acceleration(self, preview_data):
        try:
            template_path = os.path.join(self.templates_dir, "ft-nm-000-008.xlsx")
            if not os.path.exists(template_path):
                return False, f"Plantilla no encontrada: {template_path}"
                
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # --- MAPPING DATA ---
            moto = preview_data.get('moto_info', {})
            env_cond = preview_data.get('env_conditions', {})
            lugar = env_cond.get('lugar', {}) if env_cond else {}
            inputs = preview_data.get('inputs', [{}])[0]
            
            # Header info
            ws['D7'] = pd.Timestamp.now().strftime("%d/%m/%Y")
            ws['D8'] = moto.get('Placa', '')
            ws['D9'] = moto.get('Peso (Kg)', '')
            
            ws['G7'] = moto.get('Nombre Comercial', '')
            ws['G8'] = moto.get('Chasis', '')
            ws['G9'] = moto.get('Potencia (Hp)', '')
            
            ws['J7'] = moto.get('Código Modelo', '')
            ws['J8'] = moto.get('Motor', '')
            ws['J9'] = moto.get('Torque (Nm)', '')
            
            ws['A12'] = preview_data.get('comments', '')
            
            # Pilot Info
            ws['C18'] = inputs.get('pilot', '')
            ws['C19'] = inputs.get('weight', '')
            ws['C20'] = inputs.get('altura', '')
            
            # Env Cond
            ws['H18'] = lugar.get('Nombre', '')
            ctx = preview_data.get('contexto_gps', {})
            
            if ctx:
                ws['H19'] = ctx.get('altitud_promedio_msnm', lugar.get('Altitud (m)', ''))
                ws['H20'] = f"{ctx.get('latitud_inicial', '')}, {ctx.get('longitud_inicial', '')}"
                ws['H21'] = ctx.get('google_maps_link', '')
            else:
                ws['H19'] = lugar.get('Altitud (m)', '')
            
            ws['K18'] = env_cond.get('temp_amb', '') if env_cond else ''
            ws['K19'] = env_cond.get('humidity', '') if env_cond else ''
            ws['K20'] = env_cond.get('temp_ground', '') if env_cond else ''
            
            # --- RESULTS (Tables) ---
            top_3 = preview_data.get('top_3_events', [])
            for i, ev in enumerate(top_3):
                row = 52 + i
                m = ev['metrics']
                ws[f'C{row}'] = f"Evento {ev['id']}"
                ws[f'F{row}'] = m.get('v_start', 0.0)
                ws[f'G{row}'] = m.get('v_final', 80.0)
                ws[f'H{row}'] = m.get('time_s', 0.0)
                ws[f'I{row}'] = m.get('dist_m', 0.0)
                ws[f'J{row}'] = m.get('avg_acc', 0.0)
                ws[f'K{row}'] = m.get('top_rpm', 0.0)
                
            # Best event segments
            segments = preview_data.get('best_event_segments', [])
            # Segments will be [("0-20", t, d, a, rpm), ("0-40", ...)]
            for i, seg in enumerate(segments):
                row = 90 + i
                ws[f'F{row}'] = seg[1] # Tiempo
                ws[f'G{row}'] = seg[2] # Distancia
                ws[f'H{row}'] = seg[3] # Aceleracion
                ws[f'I{row}'] = seg[4] # RPM
                
            # --- IMAGES ---
            def insert_img(bytes_data, cell_or_anchor_func):
                if bytes_data:
                    if isinstance(bytes_data, dict): bytes_data = bytes_data.get('bytes')
                    try:
                        # Load via PIL to verify and optionally resize/format
                        pil_img = Image.open(io.BytesIO(bytes_data))
                        img_byte_arr = io.BytesIO()
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        
                        xl_img = OpenpyxlImage(img_byte_arr)
                        
                        if callable(cell_or_anchor_func):
                            # Calculate dimensions based on original image
                            xl_img.anchor = cell_or_anchor_func(pil_img.size)
                            ws.add_image(xl_img)
                        else:
                            ws.add_image(xl_img, cell_or_anchor_func)
                    except Exception as e:
                        print(f"Error inserting image: {e}")
                        
            def context_map_anchor(original_size):
                from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
                from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
                
                orig_w, orig_h = original_size
                width_inch = 9.97
                height_inch = width_inch * (orig_h / orig_w)
                
                emu_per_inch = 914400
                left_emu = int(6.27 * emu_per_inch)
                top_emu = int(5.69 * emu_per_inch)
                w_emu = int(width_inch * emu_per_inch)
                h_emu = int(height_inch * emu_per_inch)
                
                pos = XDRPoint2D(x=left_emu, y=top_emu)
                ext = XDRPositiveSize2D(cx=w_emu, cy=h_emu)
                return AbsoluteAnchor(pos=pos, ext=ext)
            
            insert_img(preview_data.get('context_map'), context_map_anchor)
            insert_img(preview_data.get('img_combined'), 'B56')
            insert_img(preview_data.get('img_detail_gps'), 'B95')
            insert_img(preview_data.get('img_detail_v'), 'B135')
            insert_img(preview_data.get('img_detail_a'), 'B158')
            insert_img(preview_data.get('img_detail_rpm'), 'B173')

            # --- SAVE ---
            
            def clean(s): return "".join([c for c in str(s) if c.isalnum() or c in (' ', '-', '_')]).strip()
            moto_str = clean(moto.get('Nombre Comercial', 'Moto'))
            fecha_str = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Aceleracion_{moto_str}_{fecha_str}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            wb.save(filepath)
            return True, filepath
            
        except Exception as e:
            return False, str(e)

    def generate_top_speed(self, preview_data):
        try:
            template_path = os.path.join(self.templates_dir, "ft-nm-000-007.xlsx")
            if not os.path.exists(template_path):
                return False, f"Plantilla no encontrada: {template_path}"
                
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # --- MAPPING DATA ---
            moto = preview_data.get('moto_info', {})
            env_cond = preview_data.get('env_conditions', {})
            lugar = env_cond.get('lugar', {}) if env_cond else {}
            inputs = preview_data.get('inputs', [{}])[0]
            
            # Header info
            ws['D7'] = pd.Timestamp.now().strftime("%d/%m/%Y")
            ws['D8'] = moto.get('Placa', '')
            ws['D9'] = moto.get('Peso (Kg)', '')
            
            ws['G7'] = moto.get('Nombre Comercial', '')
            ws['G8'] = moto.get('Chasis', '')
            ws['G9'] = moto.get('Potencia (Hp)', '')
            
            ws['J7'] = moto.get('Código Modelo', '')
            ws['J8'] = moto.get('Motor', '')
            ws['J9'] = moto.get('Torque (Nm)', '')
            
            ws['A12'] = preview_data.get('comments', '')
            
            # Pilot Info
            ws['C18'] = inputs.get('pilot', '')
            ws['C19'] = inputs.get('weight', '')
            ws['C20'] = inputs.get('altura', '')
            
            # Env Cond
            ws['H18'] = lugar.get('Nombre', '')
            ctx = preview_data.get('contexto_gps', {})
            
            if ctx:
                ws['H19'] = ctx.get('altitud_promedio_msnm', lugar.get('Altitud (m)', ''))
                ws['H20'] = f"{ctx.get('latitud_inicial', '')}, {ctx.get('longitud_inicial', '')}"
                ws['H21'] = ctx.get('google_maps_link', '')
            else:
                ws['H19'] = lugar.get('Altitud (m)', '')
            
            ws['K18'] = env_cond.get('temp_amb', '') if env_cond else ''
            ws['K19'] = env_cond.get('humidity', '') if env_cond else ''
            ws['K20'] = env_cond.get('temp_ground', '') if env_cond else ''
            
            # --- RESULTS (Tables) ---
            top_events = preview_data.get('top_events', [])
            for i, ev in enumerate(top_events):
                row = 52 + i
                m = ev['metrics']
                ws[f'C{row}'] = f"Evento {ev['id']}"
                ws[f'F{row}'] = m.get('v_start', 0.0)
                ws[f'G{row}'] = m.get('v_final', 0.0)
                ws[f'H{row}'] = m.get('time_s', 0.0)
                ws[f'I{row}'] = m.get('dist_m', 0.0)
                ws[f'J{row}'] = m.get('avg_acc', 0.0)
                ws[f'K{row}'] = m.get('top_rpm', 0.0)
                
            # Best event specific table
            if top_events:
                best = top_events[0]['metrics']
                ws['D90'] = best.get('v_max', 0.0)
                ws['I90'] = best.get('avg_acc', 0.0)
                ws['J90'] = best.get('top_rpm', 0.0)
                
            # --- IMAGES ---
            def insert_img(bytes_data, cell_or_anchor_func):
                if bytes_data:
                    if isinstance(bytes_data, dict): bytes_data = bytes_data.get('bytes')
                    try:
                        pil_img = Image.open(io.BytesIO(bytes_data))
                        img_byte_arr = io.BytesIO()
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        
                        xl_img = OpenpyxlImage(img_byte_arr)
                        
                        if callable(cell_or_anchor_func):
                            xl_img.anchor = cell_or_anchor_func(pil_img.size)
                            ws.add_image(xl_img)
                        else:
                            ws.add_image(xl_img, cell_or_anchor_func)
                    except Exception as e:
                        print(f"Error inserting image: {e}")
                        
            def context_map_anchor(original_size):
                from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
                from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
                
                orig_w, orig_h = original_size
                width_inch = 9.97
                height_inch = width_inch * (orig_h / orig_w)
                
                emu_per_inch = 914400
                left_emu = int(6.27 * emu_per_inch)
                top_emu = int(5.69 * emu_per_inch)
                w_emu = int(width_inch * emu_per_inch)
                h_emu = int(height_inch * emu_per_inch)
                
                pos = XDRPoint2D(x=left_emu, y=top_emu)
                ext = XDRPositiveSize2D(cx=w_emu, cy=h_emu)
                return AbsoluteAnchor(pos=pos, ext=ext)
            
            insert_img(preview_data.get('context_map'), context_map_anchor)
            insert_img(preview_data.get('img_combined'), 'C56')
            insert_img(preview_data.get('img_detail_gps'), 'B92')
            insert_img(preview_data.get('img_detail_v'), 'B132')
            insert_img(preview_data.get('img_detail_a'), 'B162')
            insert_img(preview_data.get('img_detail_rpm'), 'B176')

            # --- SAVE ---
            def clean(s): return "".join([c for c in str(s) if c.isalnum() or c in (' ', '-', '_')]).strip()
            moto_str = clean(moto.get('Nombre Comercial', 'Moto'))
            fecha_str = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            filename = f"VelMaxima_{moto_str}_{fecha_str}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            wb.save(filepath)
            return True, filepath
            
        except Exception as e:
            return False, str(e)


    def generate_accel_recovery(self, preview_data):
        try:
            template_path = os.path.join(self.templates_dir, "ft-nm-000-008.xlsx")
            if not os.path.exists(template_path):
                return False, f"Plantilla no encontrada: {template_path}"
                
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            def fmt(val, dec=2):
                if val is None or val == "": return ""
                try:
                    s = f"{float(val):.{dec}f}"
                    return s.replace('.', ',')
                except:
                    return str(val)
                    
            # --- MAPPING DATA ENCABEZADOS ---
            moto = preview_data.get('moto_info', {})
            env_cond = preview_data.get('env_conditions', {})
            lugar = env_cond.get('lugar', {}) if env_cond else {}
            inputs = preview_data.get('inputs', [{}])[0]
            ctx = preview_data.get('contexto_gps', {})
            
            ws['C7'] = pd.Timestamp.now().strftime("%d/%m/%Y")
            ws['C8'] = moto.get('Placa', '')
            ws['C9'] = fmt(moto.get('Peso (Kg)', ''), 0)
            
            ws['G7'] = moto.get('Nombre Comercial', '')
            ws['G8'] = moto.get('Chasis', '')
            ws['G9'] = fmt(moto.get('Potencia (Hp)', ''), 1)
            
            ws['K7'] = moto.get('Código Modelo', '')
            ws['K8'] = moto.get('Motor', '')
            ws['K9'] = fmt(moto.get('Torque (Nm)', ''), 1)
            ws['A12'] = preview_data.get('comments', '')
            
            ws['C18'] = inputs.get('pilot', '')
            ws['C19'] = fmt(inputs.get('weight', ''), 0)
            ws['C20'] = fmt(inputs.get('altura', ''), 0)
            
            ws['I18'] = lugar.get('Nombre', '')
            if ctx:
                ws['I19'] = fmt(ctx.get('altitud_promedio_msnm', lugar.get('Altitud (m)', '')))
                ws['I20'] = f"{ctx.get('latitud_inicial', '')}, {ctx.get('longitud_inicial', '')}"
                ws['I21'] = ctx.get('google_maps_link', '')
            else:
                ws['I19'] = fmt(lugar.get('Altitud (m)', ''))
                
            ws['L18'] = fmt(env_cond.get('temp_amb', '')) if env_cond else ''
            ws['L19'] = fmt(env_cond.get('humidity', '')) if env_cond else ''
            ws['L20'] = fmt(env_cond.get('temp_ground', '')) if env_cond else ''

            # --- IMAGE INSERTER ---
            def insert_img(bytes_data, cell_or_anchor_func):
                if bytes_data:
                    if isinstance(bytes_data, dict): bytes_data = bytes_data.get('bytes')
                    try:
                        from PIL import Image
                        from openpyxl.drawing.image import Image as OpenpyxlImage
                        pil_img = Image.open(io.BytesIO(bytes_data))
                        img_byte_arr = io.BytesIO()
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        xl_img = OpenpyxlImage(img_byte_arr)
                        if callable(cell_or_anchor_func):
                            xl_img.anchor = cell_or_anchor_func(pil_img.size)
                            ws.add_image(xl_img)
                        else:
                            ws.add_image(xl_img, cell_or_anchor_func)
                    except Exception as e:
                        print(f"Error inserting image: {e}")
                        
            def abs_anchor(orig_size, w_inch, left_inch, top_inch):
                from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
                from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
                orig_w, orig_h = orig_size
                h_inch = w_inch * (orig_h / orig_w)
                emu_per_inch = 914400
                pos = XDRPoint2D(x=int(left_inch * emu_per_inch), y=int(top_inch * emu_per_inch))
                ext = XDRPositiveSize2D(cx=int(w_inch * emu_per_inch), cy=int(h_inch * emu_per_inch))
                return AbsoluteAnchor(pos=pos, ext=ext)

            # Context Map
            insert_img(preview_data.get('context_map'), lambda s: abs_anchor(s, 9.97, 7.52, 5.69))

            # --- ACCELERATION 0-80 ---
            a_data = preview_data.get('accel_data')
            if a_data:
                # Tabla resumen mejores eventos -> H53
                for i, ev in enumerate(a_data.get('top_3_events', [])):
                    row = 53 + i
                    m = ev['metrics']
                    ws[f'H{row}'] = f"Evento {ev['id']}"
                    ws[f'I{row}'] = fmt(m.get('v_start', 0))
                    ws[f'J{row}'] = fmt(m.get('v_final', 0))
                    ws[f'K{row}'] = fmt(m.get('time_s', 0))
                    ws[f'L{row}'] = fmt(m.get('dist_m', 0))
                    ws[f'M{row}'] = fmt(m.get('avg_acc', 0))
                    ws[f'N{row}'] = fmt(m.get('top_rpm', 0), 0)
                
                # Image
                insert_img(a_data.get('img_combined'), lambda s: abs_anchor(s, 8.7, 8.78, 13.88))
                
                # Segments J81 to M85
                segs = a_data.get('segments', [])
                # segs has format: ["0-20", "time", "dist", "acc", "rpm"] (5 items)
                # expected rows: 81 (0-20), 82 (20-40), 83 (40-60), 84 (60-80), 85 (0-80).
                # The segs list is guaranteed to have the right order if present.
                for i, seg in enumerate(segs):
                    if i > 4: break
                    row = 81 + i
                    ws[f'J{row}'] = fmt(seg[1])
                    ws[f'K{row}'] = fmt(seg[2])
                    ws[f'L{row}'] = fmt(seg[3])
                    ws[f'M{row}'] = fmt(seg[4], 0)
                    
                # Images detailed accel
                insert_img(a_data.get('img_detail_gps'), lambda s: abs_anchor(s, 8.64, 8.81, 21.38))
                insert_img(a_data.get('img_detail_v'), lambda s: abs_anchor(s, 8.64, 8.81, 27.75))
                insert_img(a_data.get('img_detail_a'), lambda s: abs_anchor(s, 8.64, 8.81, 31.67))
                insert_img(a_data.get('img_detail_rpm'), lambda s: abs_anchor(s, 8.64, 8.81, 33.96))

            # --- RECOVERY ---
            r_data = preview_data.get('recovery_data')
            if r_data:
                # Tabla resumen: A53
                for i, ev in enumerate(r_data.get('summary_events', [])):
                    row = 53 + i
                    m = ev['metrics']
                    # Use the speed group as identifier since it's mixed:
                    ws[f'A{row}'] = f"Evento {ev['id']} ({int(m.get('v_start',0))}-80)"
                    ws[f'B{row}'] = fmt(m.get('v_start', 0))
                    ws[f'C{row}'] = fmt(m.get('v_final', 0))
                    ws[f'D{row}'] = fmt(m.get('time_s', 0))
                    ws[f'E{row}'] = fmt(m.get('dist_m', 0))
                    ws[f'F{row}'] = fmt(m.get('avg_acc', 0))
                    ws[f'G{row}'] = fmt(m.get('top_rpm', 0), 0)

                insert_img(r_data.get('summary_img'), lambda s: abs_anchor(s, 8.7, 0.03, 13.88))
                
                bands_map = {
                    30: {"row": 81, "gps_top": 20.4, "v_top": 27.0, "a_top": 31.04, "rpm_top": 33.39},
                    40: {"row": 147, "gps_top": 37.5, "v_top": 43.5, "a_top": 47.5, "rpm_top": 50.0},
                    50: {"row": 213, "gps_top": 53.5, "v_top": 60.0, "a_top": 64.0, "rpm_top": 66.5}
                }
                
                bands = r_data.get('bands', {})
                for spd, conf in bands_map.items():
                    b_info = bands.get(spd)
                    if b_info:
                        m = b_info['best_event']['metrics']
                        rw = conf['row']
                        ws[f'B{rw}'] = fmt(m.get('v_start', 0))
                        ws[f'C{rw}'] = fmt(m.get('v_final', 0))
                        ws[f'D{rw}'] = fmt(m.get('time_s', 0))
                        ws[f'E{rw}'] = fmt(m.get('dist_m', 0))
                        ws[f'F{rw}'] = fmt(m.get('avg_acc', 0))
                        ws[f'G{rw}'] = fmt(m.get('top_rpm', 0), 0)
                        
                        insert_img(b_info.get('img_gps'), lambda s: abs_anchor(s, 8.64, 0.06, conf['gps_top']))
                        insert_img(b_info.get('img_v'), lambda s: abs_anchor(s, 8.64, 0.06, conf['v_top']))
                        insert_img(b_info.get('img_a'), lambda s: abs_anchor(s, 8.64, 0.06, conf['a_top']))
                        insert_img(b_info.get('img_rpm'), lambda s: abs_anchor(s, 8.64, 0.06, conf['rpm_top']))

            # --- SAVE ---
            def cln(s): return "".join([c for c in str(s) if c.isalnum() or c in (' ', '-', '_')]).strip()
            moto_str = cln(moto.get('Nombre Comercial', 'Moto'))
            fecha_str = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Aceleracion_y_Recuperacion_{moto_str}_{fecha_str}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            wb.save(filepath)
            return True, filepath
            
        except Exception as e:
            return False, str(e)
    